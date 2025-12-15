#!/usr/bin/env python3
"""
简易发票 OCR 汇总工具

特点：
- 递归扫描当前目录下的 PDF/图片发票
- PDF 首页通过 `pdftoppm` 转成 PNG 后送入 Ollama
- 调用局域网 Ollama qwen3-vl:8b 提取完整发票信息
- 支持多维度统计分析和文件重命名
- ✅ 自动跳过含有"行程单"的文件
- ✅ 自动验证发票，跳过非发票文件
- ✅ 修复长文件名导致的 PDF 处理失败问题

使用：
  python3 invoice_ocr_sum.py                    # 扫描当前目录
  python3 invoice_ocr_sum.py /path/to/dir      # 扫描指定目录
  python3 invoice_ocr_sum.py --excel --rename  # 生成报告和重命名
"""

from __future__ import annotations

import argparse
import base64
import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile
import time
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Tuple, Dict, Optional
from urllib.error import URLError, HTTPError
from urllib.request import Request, urlopen

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class InvoiceInfo:
    """发票信息数据类"""
    # 基本信息
    invoice_no: str = ""
    issue_date: str = ""
    seller: str = ""
    buyer: str = ""
    total: float = 0.0
    tax: float = 0.0
    subtotal: float = 0.0
    items: str = ""
    notes: str = ""
    
    # 发票分类（新增）
    invoice_type: str = ""  # special_vat, general_vat, electronic, toll, taxi, train, flight, other
    invoice_type_name: str = ""  # 类型中文名
    expense_category: str = ""  # travel, dining, office, transport, telecom, conference, training, service, material, other
    expense_category_name: str = ""  # 类别中文名
    
    # 真伪验证（新增）
    risk_level: str = ""  # low, medium, high
    risk_notes: str = ""  # 风险说明
    has_stamp: bool = True  # 是否有印章
    image_quality: str = ""  # good, fair, poor


DEFAULT_PROMPT = (
    "你是发票识别专家。请仔细识别图片中的发票，按JSON格式返回数据。\n"
    "\n"
    "🔴【最重要】价税合计金额（total）必须准确识别！\n"
    "这是最核心的字段，仔细查找发票上的「价税合计」或「合计」或「总金额」。\n"
    "\n"
    "字段说明：\n"
    "• total: 价税合计（最重要！查找发票最下方的合计金额。仅数字，如1234.56）\n"
    "• invoice_no: 发票号码（如00123456）\n"
    "• issue_date: 开票日期（YYYY-MM-DD，如2024-12-01）\n"
    "• buyer: 购买方/买方名称（需要准确）\n"
    "• seller: 供应商/卖方名称\n"
    "• tax: 税额（仅数字，无则为0）\n"
    "• subtotal: 小计（仅数字，无则为0）\n"
    "• items: 商品/服务项目（逗号分隔，最多3个）\n"
    "• notes: 备注（空即可）\n"
    "\n"
    "返回格式（仅返回JSON，无其他内容）：\n"
    "{\n"
    '  "invoice_no": "",\n'
    '  "issue_date": "YYYY-MM-DD",\n'
    '  "seller": "",\n'
    '  "buyer": "",\n'
    '  "total": 0,\n'
    '  "tax": 0,\n'
    '  "subtotal": 0,\n'
    '  "items": "",\n'
    '  "notes": ""\n'
    "}\n"
    "\n"
    "⚠️ 特别提醒：\n"
    "1. total 是最关键的字段，必须准确（宁可留空也不要错误的金额）\n"
    "2. 如果不是发票，返回所有字段为空或0\n"
    "3. 如某字段无法识别，返回空字符串或0，不要猜测\n"
    "4. 只返回JSON，不要添加任何说明文字"
)

VALIDATE_PROMPT = (
    "请判断图片中的文件是否是发票。\n"
    "如果是发票（增值税发票、普通发票等），返回 {\"is_invoice\": true}\n"
    "如果不是发票（行程单、收据等），返回 {\"is_invoice\": false}\n"
    "不要输出其他任何内容。"
)

# 发票真伪验证提示词
VERIFY_INVOICE_PROMPT = (
    "你是发票审核专家。请仔细检查这张发票的真实性和完整性。\n"
    "\n"
    "请检查以下项目并按JSON格式返回：\n"
    "1. 发票印章是否清晰可见\n"
    "2. 发票代码和发票号码是否完整\n"
    "3. 密码区/校验码是否存在（增值税发票）\n"
    "4. 二维码是否存在（电子发票）\n"
    "5. 图片质量是否清晰、完整\n"
    "6. 是否有明显的修改/PS痕迹\n"
    "7. 金额数字与大写是否一致\n"
    "\n"
    "返回格式（仅JSON）：\n"
    "{\n"
    '  "risk_level": "low/medium/high",\n'
    '  "has_stamp": true/false,\n'
    '  "has_complete_code": true/false,\n'
    '  "has_qrcode": true/false,\n'
    '  "image_quality": "good/fair/poor",\n'
    '  "has_tampering": true/false,\n'
    '  "amount_consistent": true/false,\n'
    '  "risk_notes": "具体问题描述（如有）"\n'
    "}\n"
    "\n"
    "风险等级判断标准：\n"
    "- low: 发票完整、清晰、无异常\n"
    "- medium: 存在轻微问题（如图片略模糊、部分信息不清晰）\n"
    "- high: 存在严重问题（无印章、有修改痕迹、金额不一致等）"
)

# 发票分类提示词
CLASSIFY_INVOICE_PROMPT = (
    "请识别这张发票的类型和费用类别，按JSON格式返回。\n"
    "\n"
    "发票类型（invoice_type）：\n"
    "- special_vat: 增值税专用发票\n"
    "- general_vat: 增值税普通发票\n"
    "- electronic: 电子发票\n"
    "- toll: 通行费发票\n"
    "- taxi: 出租车发票\n"
    "- train: 火车票\n"
    "- flight: 机票行程单\n"
    "- other: 其他类型\n"
    "\n"
    "费用类别（expense_category）：\n"
    "- travel: 差旅\n"
    "- dining: 餐饮\n"
    "- office: 办公用品\n"
    "- transport: 交通\n"
    "- telecom: 通讯\n"
    "- conference: 会议\n"
    "- training: 培训\n"
    "- service: 服务费\n"
    "- material: 材料/设备\n"
    "- other: 其他\n"
    "\n"
    "返回格式（仅JSON）：\n"
    "{\n"
    '  "invoice_type": "类型代码",\n'
    '  "invoice_type_name": "类型中文名",\n'
    '  "expense_category": "类别代码",\n'
    '  "expense_category_name": "类别中文名"\n'
    "}"
)

OLLAMA_HOST = "192.168.110.219"
OLLAMA_PORT = 11434
OLLAMA_MODEL = "qwen3-vl:8b"

# 统一 OCR Provider（由 GUI 设置）
OCR_PROVIDER = None


def get_pdftoppm_path() -> str:
    """获取 pdftoppm 的路径（支持打包后的应用）"""
    import platform
    is_windows = platform.system() == 'Windows'
    exe_suffix = '.exe' if is_windows else ''

    # 尝试多个可能的 pdftoppm 路径
    possible_paths = [
        # PyInstaller 打包后的路径 (bin/pdftoppm)
        os.path.join(getattr(sys, '_MEIPASS', ''), 'bin', f'pdftoppm{exe_suffix}'),
        # 如果是 macOS .app bundle
        os.path.join(os.path.dirname(sys.executable), '..', 'Frameworks', 'bin', 'pdftoppm'),
        # 系统 PATH
        f"pdftoppm{exe_suffix}",
        # Homebrew (M1/M2 Mac)
        "/opt/homebrew/bin/pdftoppm",
        # Homebrew (Intel Mac)
        "/usr/local/bin/pdftoppm",
    ]
    
    for path in possible_paths:
        if not path:
            continue
        try:
            result = subprocess.run(
                [path, "-v"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False
            )
            if result.returncode == 0 or result.returncode == 99:  # 99 是正常的版本输出码
                return path
        except Exception:
            continue
    
    return None


def check_pdftoppm() -> bool:
    """检查 pdftoppm 是否可用"""
    return get_pdftoppm_path() is not None


def run_pdftoppm_first_page(pdf_path: Path, tmpdir: Path) -> Path:
    """将 PDF 的第一页转换成 PNG，并返回图片路径（使用短标识符避免路径过长）。"""
    # 检查 pdftoppm 是否可用
    if not check_pdftoppm():
        raise RuntimeError(
            "未找到 pdftoppm 工具。请安装 poppler:\n"
            "  macOS: brew install poppler\n"
            "  如果使用打包的应用，PDF 文件暂不支持，请转换为图片格式。"
        )
    
    # 使用短标识符避免路径过长
    short_id = hashlib.md5(str(pdf_path).encode()).hexdigest()[:8]
    output_prefix = tmpdir / short_id

    # 获取 pdftoppm 路径
    pdftoppm = get_pdftoppm_path()
    if not pdftoppm:
        raise RuntimeError(
            "未找到 pdftoppm 工具。请安装 poppler:\n"
            "  macOS: brew install poppler\n"
            "  如果使用打包的应用，PDF 文件暂不支持，请转换为图片格式。"
        )
    
    cmd = [
        pdftoppm,
        "-png",
        "-singlefile",
        "-f",
        "1",
        "-l",
        "1",
        str(pdf_path),
        str(output_prefix),
    ]
    
    proc = subprocess.run(cmd, check=False, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE)
    if proc.returncode != 0:
        error_msg = proc.stderr.decode('utf-8', 'ignore').strip()
        raise RuntimeError(f"pdftoppm 转换失败: {error_msg}")
    
    out_png = output_prefix.with_suffix(".png")
    if not out_png.exists():
        raise FileNotFoundError(f"pdftoppm 未生成输出文件: {out_png}")
    
    return out_png


def call_ollama_ocr(
    image_path: Path,
    host: str,
    port: int,
    model: str,
    prompt: str,
    timeout: int = 300,
) -> str:
    """调用 OCR（支持统一 Provider 或 Ollama）"""
    # 优先使用统一 Provider
    if OCR_PROVIDER is not None:
        try:
            return OCR_PROVIDER.call_ocr(image_path, prompt, timeout)
        except Exception as e:
            raise RuntimeError(f"OCR API 调用失败: {e}")
    
    # 回退到原有 Ollama 调用
    with image_path.open("rb") as f:
        image_b64 = base64.b64encode(f.read()).decode("ascii")

    payload = {
        "model": model,
        "messages": [
            {
                "role": "user",
                "content": prompt,
                "images": [image_b64],
            }
        ],
        "stream": False,
    }
    url = f"http://{host}:{port}/api/chat"
    req = Request(url, data=json.dumps(payload).encode("utf-8"), headers={"Content-Type": "application/json"})
    try:
        with urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        return data.get("message", {}).get("content", "")
    except Exception as e:
        raise RuntimeError(f"Ollama API 调用失败: {e}")


def parse_invoice_info(response_text: str) -> InvoiceInfo:
    """解析模型返回的JSON文本，提取完整的发票信息。"""
    info = InvoiceInfo()

    try:
        data = json.loads(response_text)
        if isinstance(data, dict):
            info.invoice_no = str(data.get("invoice_no", "")).strip()
            info.issue_date = str(data.get("issue_date", "")).strip()
            info.seller = str(data.get("seller", "")).strip()
            info.buyer = str(data.get("buyer", "")).strip()
            info.items = str(data.get("items", "")).strip()
            info.notes = str(data.get("notes", "")).strip()

            # 解析数值字段
            for field in ["total", "tax", "subtotal"]:
                val = data.get(field, 0)
                if isinstance(val, (int, float)):
                    setattr(info, field, float(val))
                elif isinstance(val, str):
                    try:
                        setattr(info, field, float(val.replace(",", "").strip() or 0))
                    except (ValueError, AttributeError):
                        setattr(info, field, 0.0)
    except Exception:
        pass

    return info


def validate_is_invoice(image_path: Path, args) -> bool:
    """验证文件是否是发票（可选，避免处理非发票文件）。"""
    try:
        response = call_ollama_ocr(image_path, args.host, args.port, args.model, VALIDATE_PROMPT, timeout=60)
        data = json.loads(response)
        return data.get("is_invoice", False)
    except Exception:
        # 验证失败时假定为发票，继续处理
        return True


def verify_invoice(image_path: Path, args) -> dict:
    """验证发票真伪和完整性。
    
    返回包含风险等级和详细信息的字典。
    """
    default_result = {
        "risk_level": "low",
        "has_stamp": True,
        "has_complete_code": True,
        "has_qrcode": False,
        "image_quality": "good",
        "has_tampering": False,
        "amount_consistent": True,
        "risk_notes": ""
    }
    
    try:
        response = call_ollama_ocr(image_path, args.host, args.port, args.model, VERIFY_INVOICE_PROMPT, timeout=90)
        data = json.loads(response)
        if isinstance(data, dict):
            return {
                "risk_level": data.get("risk_level", "low"),
                "has_stamp": data.get("has_stamp", True),
                "has_complete_code": data.get("has_complete_code", True),
                "has_qrcode": data.get("has_qrcode", False),
                "image_quality": data.get("image_quality", "good"),
                "has_tampering": data.get("has_tampering", False),
                "amount_consistent": data.get("amount_consistent", True),
                "risk_notes": data.get("risk_notes", "")
            }
    except Exception as e:
        default_result["risk_notes"] = f"验证失败: {str(e)[:50]}"
    
    return default_result


def classify_invoice(image_path: Path, args) -> dict:
    """识别发票类型和费用类别。
    
    返回包含发票类型和费用类别的字典。
    """
    default_result = {
        "invoice_type": "other",
        "invoice_type_name": "其他类型",
        "expense_category": "other",
        "expense_category_name": "其他"
    }
    
    try:
        response = call_ollama_ocr(image_path, args.host, args.port, args.model, CLASSIFY_INVOICE_PROMPT, timeout=60)
        data = json.loads(response)
        if isinstance(data, dict):
            return {
                "invoice_type": data.get("invoice_type", "other"),
                "invoice_type_name": data.get("invoice_type_name", "其他类型"),
                "expense_category": data.get("expense_category", "other"),
                "expense_category_name": data.get("expense_category_name", "其他")
            }
    except Exception:
        pass
    
    return default_result


def iter_invoice_files(root: Path) -> Iterable[Path]:
    """递归扫描发票文件（跳过行程单和非发票文件）。"""
    exts = {".pdf", ".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}
    skip_keywords = {"行程单", "itinerary", "receipt"}  # 跳过的关键词

    for path in root.rglob("*"):
        if path.is_file() and path.suffix.lower() in exts:
            # 跳过含有特定关键词的文件
            if any(keyword in path.name.lower() for keyword in skip_keywords):
                continue
            yield path


def process_file(path: Path, args, max_retries: int = 3) -> Tuple[InvoiceInfo, List[str]]:
    """处理单个文件，支持失败重试。"""
    errors: List[str] = []
    info = InvoiceInfo()
    retry_count = 0

    while retry_count <= max_retries:
        try:
            if path.suffix.lower() == ".pdf":
                with tempfile.TemporaryDirectory(prefix="invoice_ocr_") as tmp:
                    png = run_pdftoppm_first_page(path, Path(tmp))

                    # 验证是否为发票（只验证一次，不重试）
                    if retry_count == 0 and not validate_is_invoice(png, args):
                        errors.append("非发票")
                        return info, errors

                    try:
                        response = call_ollama_ocr(png, args.host, args.port, args.model, args.prompt)
                        info = parse_invoice_info(response)
                        # 如果成功识别到金额，返回结果
                        if info.total > 0:
                            return info, errors
                        # 如果没有识别到金额，继续重试
                        if retry_count < max_retries:
                            retry_count += 1
                            time.sleep(2)  # 等待2秒后重试
                            continue
                        else:
                            errors.append("OCR 失败: 未识别到金额")
                            return info, errors
                    except (HTTPError, URLError) as e:
                        if retry_count < max_retries:
                            retry_count += 1
                            time.sleep(3)  # 网络错误等待3秒
                            continue
                        errors.append(f"Ollama 网络错误: {e}")
                        return info, errors
                    except Exception as e:
                        if retry_count < max_retries:
                            retry_count += 1
                            time.sleep(2)
                            continue
                        errors.append(f"OCR 失败: {e}")
                        return info, errors
            else:
                # 验证是否为发票（只验证一次，不重试）
                if retry_count == 0 and not validate_is_invoice(path, args):
                    errors.append("非发票")
                    return info, errors

                try:
                    response = call_ollama_ocr(path, args.host, args.port, args.model, args.prompt)
                    info = parse_invoice_info(response)
                    # 如果成功识别到金额，返回结果
                    if info.total > 0:
                        return info, errors
                    # 如果没有识别到金额，继续重试
                    if retry_count < max_retries:
                        retry_count += 1
                        time.sleep(2)
                        continue
                    else:
                        errors.append("OCR 失败: 未识别到金额")
                        return info, errors
                except (HTTPError, URLError) as e:
                    if retry_count < max_retries:
                        retry_count += 1
                        time.sleep(3)
                        continue
                    errors.append(f"Ollama 网络错误: {e}")
                    return info, errors
                except Exception as e:
                    if retry_count < max_retries:
                        retry_count += 1
                        time.sleep(2)
                        continue
                    errors.append(f"OCR 失败: {e}")
                    return info, errors
        except Exception as e:
            if retry_count < max_retries:
                retry_count += 1
                time.sleep(2)
                continue
            errors.append(f"预处理失败: {e}")
            return info, errors

    return info, errors


def validate_and_analyze(invoices: List[Tuple[Path, InvoiceInfo, List[str]]]) -> Dict:
    """数据验证和分析。"""
    analysis = {
        "total_count": len(invoices),
        "valid_count": sum(1 for _, info, errs in invoices if info.total > 0 and not errs),
        "total_amount": sum(info.total for _, info, _ in invoices),
        "duplicates": [],
        "warnings": [],
        "by_month": {},
        "by_seller": {},
        "by_amount_range": {"0-1000": 0, "1000-10000": 0, "10000+": 0},
    }

    # 检查重复发票号
    invoice_nos = {}
    for path, info, _ in invoices:
        if info.invoice_no:
            if info.invoice_no in invoice_nos:
                analysis["duplicates"].append(info.invoice_no)
            invoice_nos[info.invoice_no] = path

    # 按月份统计
    for path, info, _ in invoices:
        if info.issue_date:
            try:
                month = info.issue_date[:7]  # YYYY-MM
                if month not in analysis["by_month"]:
                    analysis["by_month"][month] = {"count": 0, "total": 0.0}
                analysis["by_month"][month]["count"] += 1
                analysis["by_month"][month]["total"] += info.total
            except Exception:
                pass

    # 按供应商统计
    for path, info, _ in invoices:
        if info.seller:
            seller = info.seller[:20]  # 截断长名称
            if seller not in analysis["by_seller"]:
                analysis["by_seller"][seller] = {"count": 0, "total": 0.0}
            analysis["by_seller"][seller]["count"] += 1
            analysis["by_seller"][seller]["total"] += info.total

    # 按金额区间统计
    for path, info, _ in invoices:
        if info.total > 0:
            if info.total < 1000:
                analysis["by_amount_range"]["0-1000"] += 1
            elif info.total < 10000:
                analysis["by_amount_range"]["1000-10000"] += 1
            else:
                analysis["by_amount_range"]["10000+"] += 1

    # 异常检测
    if analysis["total_count"] > 0:
        avg_amount = analysis["total_amount"] / analysis["total_count"]
        for path, info, errs in invoices:
            if info.total > 0 and info.total > avg_amount * 3:
                analysis["warnings"].append(f"{path.name}: 金额 {info.total:.2f} 元（超过平均值3倍）")

    return analysis


def generate_excel_report(
    invoices: List[Tuple[Path, InvoiceInfo, List[str]]],
    analysis: Dict,
    output_path: Path
) -> bool:
    """生成 Excel 详细报告。"""
    if not HAS_OPENPYXL:
        return False

    try:
        from openpyxl import Workbook
        wb = Workbook()

        # 工作表1: 详细清单
        ws_detail = wb.active
        ws_detail.title = "发票明细"

        # 检查是否有验证/分类数据
        has_verify = any(info.risk_level for _, info, _ in invoices)
        has_classify = any(info.invoice_type for _, info, _ in invoices)
        
        # 动态構建表头
        headers = ["序号", "文件名", "发票号", "开票日期", "供应商", "购买方", "合计金额", "税额", "小计"]
        if has_classify:
            headers.extend(["发票类型", "费用类别"])
        if has_verify:
            headers.extend(["风险等级", "风险说明"])
        headers.extend(["项目", "状态"])
        
        ws_detail.append(headers)

        # 样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # 风险等级颜色
        risk_fills = {
            "high": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            "medium": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
            "low": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
        }

        for cell in ws_detail[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        for idx, (path, info, errors) in enumerate(invoices, 1):
            row = [
                idx,
                path.name,
                info.invoice_no,
                info.issue_date,
                info.seller[:30] if info.seller else "",
                info.buyer[:30] if info.buyer else "",
                info.total,
                info.tax,
                info.subtotal,
            ]
            if has_classify:
                row.extend([info.invoice_type_name or "", info.expense_category_name or ""])
            if has_verify:
                risk_label = {"": "", "low": "✅ 低风险", "medium": "⚠️ 中风险", "high": "❌ 高风险"}.get(info.risk_level, "")
                row.extend([risk_label, info.risk_notes or ""])
            row.extend([
                info.items[:40] if info.items else "",
                "❌ " + errors[0][:30] if errors else "✓ OK",
            ])
            ws_detail.append(row)
            
            # 为风险等级单元格添加颜色
            if has_verify and info.risk_level in risk_fills:
                risk_col = 10 + (2 if has_classify else 0)  # 风险等级列
                ws_detail.cell(row=idx + 1, column=risk_col).fill = risk_fills[info.risk_level]

        # 列宽 (使用openpyxl工具函数支持超过26列)
        from openpyxl.utils import get_column_letter
        col_widths = [8, 25, 15, 12, 20, 20, 12, 12, 12]
        if has_classify:
            col_widths.extend([15, 12])
        if has_verify:
            col_widths.extend([12, 30])
        col_widths.extend([30, 30])
        
        for i, width in enumerate(col_widths, 1):
            ws_detail.column_dimensions[get_column_letter(i)].width = width

        # 数字格式
        for row in ws_detail.iter_rows(min_row=2, max_row=len(invoices) + 1, min_col=7, max_col=9):
            for cell in row:
                cell.number_format = "0.00"
                cell.border = border

        # 工作表2: 统计汇总
        ws_summary = wb.create_sheet("统计汇总")
        ws_summary.append(["发票统计汇总"])
        ws_summary.append([])
        ws_summary.append(["指标", "数值"])

        summary_data = [
            ["发票总数", analysis["total_count"]],
            ["有效发票数", analysis["valid_count"]],
            ["总金额", analysis["total_amount"]],
            ["平均金额", analysis["total_amount"] / max(analysis["total_count"], 1)],
            ["重复发票号", len(analysis["duplicates"])],
        ]

        for row in summary_data:
            ws_summary.append(row)

        ws_summary.append([])
        ws_summary.append(["按月份统计"])
        ws_summary.append(["月份", "数量", "合计"])
        for month, data in sorted(analysis["by_month"].items()):
            ws_summary.append([month, data["count"], data["total"]])

        ws_summary.append([])
        ws_summary.append(["按供应商统计"])
        ws_summary.append(["供应商", "数量", "合计"])
        for seller, data in sorted(analysis["by_seller"].items(), key=lambda x: x[1]["total"], reverse=True)[:10]:
            ws_summary.append([seller, data["count"], data["total"]])

        # 按发票类型统计（如果有分类数据）
        if has_classify:
            type_stats = {}
            for _, info, _ in invoices:
                if info.invoice_type_name:
                    if info.invoice_type_name not in type_stats:
                        type_stats[info.invoice_type_name] = {"count": 0, "total": 0}
                    type_stats[info.invoice_type_name]["count"] += 1
                    type_stats[info.invoice_type_name]["total"] += info.total or 0
            
            if type_stats:
                ws_summary.append([])
                ws_summary.append(["按发票类型统计"])
                ws_summary.append(["发票类型", "数量", "合计"])
                for type_name, data in sorted(type_stats.items(), key=lambda x: x[1]["count"], reverse=True):
                    ws_summary.append([type_name, data["count"], data["total"]])

        # 按费用类别统计（如果有分类数据）
        if has_classify:
            category_stats = {}
            for _, info, _ in invoices:
                if info.expense_category_name:
                    if info.expense_category_name not in category_stats:
                        category_stats[info.expense_category_name] = {"count": 0, "total": 0}
                    category_stats[info.expense_category_name]["count"] += 1
                    category_stats[info.expense_category_name]["total"] += info.total or 0
            
            if category_stats:
                ws_summary.append([])
                ws_summary.append(["按费用类别统计"])
                ws_summary.append(["费用类别", "数量", "合计"])
                for cat_name, data in sorted(category_stats.items(), key=lambda x: x[1]["count"], reverse=True):
                    ws_summary.append([cat_name, data["count"], data["total"]])

        # 按风险等级统计（如果有验证数据）
        if has_verify:
            risk_stats = {"low": 0, "medium": 0, "high": 0, "unknown": 0}
            for _, info, _ in invoices:
                if info.risk_level in risk_stats:
                    risk_stats[info.risk_level] += 1
                elif info.risk_level:
                    risk_stats["unknown"] += 1
            
            ws_summary.append([])
            ws_summary.append(["按风险等级统计"])
            ws_summary.append(["风险等级", "数量"])
            risk_labels = {"low": "✅ 低风险", "medium": "⚠️ 中风险", "high": "❌ 高风险"}
            for level in ["high", "medium", "low"]:
                if risk_stats[level] > 0:
                    ws_summary.append([risk_labels[level], risk_stats[level]])

        # 列宽
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 15
        ws_summary.column_dimensions["C"].width = 15

        wb.save(str(output_path))
        return True
    except Exception as e:
        print(f"[警告] Excel 导出失败: {e}", file=sys.stderr)
        return False


def rename_invoice_files(invoices: List[Tuple[Path, InvoiceInfo, List[str]]], rename: bool = False) -> List[str]:
    """生成文件重命名建议（格式：金额-购买方名称）。"""
    rename_ops = []
    for path, info, errors in invoices:
        if errors or not info.total or not info.buyer:
            continue

        # 生成新名称：金额-购买方名称
        buyer_short = "".join(info.buyer.split())[:15]  # 去空格，截断15字
        new_name = f"{info.total:.0f}-{buyer_short}{path.suffix}"
        new_path = path.parent / new_name

        if rename and path != new_path:
            try:
                path.rename(new_path)
                rename_ops.append(f"✓ {path.name} -> {new_name}")
            except Exception as e:
                rename_ops.append(f"✗ {path.name}: {e}")
        else:
            rename_ops.append(f"→ {path.name} -> {new_name}")

    return rename_ops


def main() -> int:
    parser = argparse.ArgumentParser(description="发票 OCR 智能汇总工具（支持多维度统计分析和文件重命名）")
    parser.add_argument("root", nargs="?", default=".", help="要扫描的目录，默认当前目录")
    parser.add_argument("--rename", action="store_true", help="启用文件重命名（金额-购买方格式）")
    parser.add_argument("--excel", action="store_true", help="生成 Excel 详细报告")
    parser.add_argument("--validate", action="store_true", help="启用非发票验证过滤")
    parser.add_argument("--max-retries", type=int, default=3, help="OCR 失败时的最大重试次数（默认3次）")
    args = parser.parse_args()

    root = Path(args.root).resolve()
    if not root.exists():
        print(f"[错误] 路径不存在: {root}", file=sys.stderr)
        return 1

    files = list(iter_invoice_files(root))
    if not files:
        print(f"[提示] 在 {root} 下未找到发票文件（PDF/图片）。")
        return 0

    # 设置 Ollama 配置
    args.host = OLLAMA_HOST
    args.port = OLLAMA_PORT
    args.model = OLLAMA_MODEL
    args.prompt = DEFAULT_PROMPT

    print(f"共发现 {len(files)} 份发票，开始 OCR ...\n")
    print(f"Ollama 地址: http://{args.host}:{args.port}  模型: {args.model}\n")

    invoices: List[Tuple[Path, InvoiceInfo, List[str]]] = []
    non_invoice_count = 0
    for idx, path in enumerate(files, 1):
        info, errors = process_file(path, args, max_retries=args.max_retries)
        if "非发票" in str(errors):
            non_invoice_count += 1
        status = "✓ OK" if not errors else f"⚠ {errors[0][:40]}"
        print(f"[{idx:03d}] {path.name:<40} -> {info.total:>10.2f} 元  {status}")
        invoices.append((path, info, errors))

    # 数据分析与验证
    print("\n" + "=" * 80)
    analysis = validate_and_analyze(invoices)

    print("📊 统计汇总")
    print(f"  发票总数：{analysis['total_count']}")
    print(f"  有效发票：{analysis['valid_count']}")
    if args.validate and non_invoice_count > 0:
        print(f"  非发票文件：{non_invoice_count}")
    print(f"  总金额：{analysis['total_amount']:.2f} 元")
    print(f"  平均金额：{analysis['total_amount'] / max(analysis['total_count'], 1):.2f} 元")

    if analysis["duplicates"]:
        print(f"\n  ⚠ 重复发票号: {', '.join(analysis['duplicates'])}")

    if analysis["warnings"]:
        print(f"\n⚠ 异常警告（超过平均值3倍）:")
        for warn in analysis["warnings"]:
            print(f"  - {warn}")

    # 按金额区间统计
    print(f"\n💰 按金额区间统计:")
    for range_key, count in analysis["by_amount_range"].items():
        print(f"  {range_key} 元: {count} 份")

    # 按月份统计（最近6个月）
    if analysis["by_month"]:
        print(f"\n📅 按月份统计:")
        for month in sorted(analysis["by_month"].keys())[-6:]:
            data = analysis["by_month"][month]
            print(f"  {month}: {data['count']} 份，合计 {data['total']:.2f} 元")

    # 按供应商统计（top 10）
    if analysis["by_seller"]:
        print(f"\n🏢 按供应商统计（top 10）:")
        for seller, data in sorted(analysis["by_seller"].items(), key=lambda x: x[1]["total"], reverse=True)[:10]:
            print(f"  {seller:<20} {data['count']:>3} 份，合计 {data['total']:>10.2f} 元")

    # 生成 Markdown 报告
    print("\n" + "=" * 80)
    output_md = root / "invoice_summary.md"
    lines = [
        "# 📋 发票 OCR 汇总报告",
        f"- 🗂️ 扫描目录：`{root}`",
        f"- 📊 发票数量：{analysis['total_count']} 份",
        f"- ✅ 有效发票：{analysis['valid_count']} 份",
        f"- 💰 总金额：**{analysis['total_amount']:.2f} 元**",
        f"- 📈 平均金额：{analysis['total_amount'] / max(analysis['total_count'], 1):.2f} 元",
        "",
        "## 📝 明细表",
        "| 序号 | 文件 | 发票号 | 日期 | 供应商 | 金额(元) | 状态 |",
        "| --- | --- | --- | --- | --- | --- | --- |",
    ]

    for i, (path, info, errors) in enumerate(invoices, 1):
        rel = path.relative_to(root)
        status = "✓" if not errors else "✗"
        lines.append(
            f"| {i} | `{rel.name}` | {info.invoice_no} | {info.issue_date} | "
            f"{info.seller[:15]} | {info.total:.2f} | {status} |"
        )

    if analysis["by_month"]:
        lines.append("")
        lines.append("## 📅 按月份统计")
        lines.append("| 月份 | 数量 | 合计(元) |")
        lines.append("| --- | --- | --- |")
        for month in sorted(analysis["by_month"].keys()):
            data = analysis["by_month"][month]
            lines.append(f"| {month} | {data['count']} | {data['total']:.2f} |")

    if analysis["by_seller"]:
        lines.append("")
        lines.append("## 🏢 按供应商统计（top 10）")
        lines.append("| 供应商 | 数量 | 合计(元) |")
        lines.append("| --- | --- | --- |")
        for seller, data in sorted(analysis["by_seller"].items(), key=lambda x: x[1]["total"], reverse=True)[:10]:
            lines.append(f"| {seller} | {data['count']} | {data['total']:.2f} |")

    try:
        output_md.write_text("\n".join(lines), encoding="utf-8")
        print(f"✅ Markdown 报告: {output_md}")
    except Exception as e:
        print(f"❌ Markdown 导出失败: {e}", file=sys.stderr)

    # 生成 Excel 报告
    if args.excel or HAS_OPENPYXL:
        output_xlsx = root / "invoice_summary.xlsx"
        if generate_excel_report(invoices, analysis, output_xlsx):
            print(f"✅ Excel 报告: {output_xlsx}")
        else:
            print("ℹ️  Excel 库未安装 (openpyxl)，跳过 Excel 导出")

    # 文件重命名建议
    if len([p for p, i, e in invoices if i.total > 0 and not e]) > 0:
        print("\n" + "=" * 80)
        print("📝 文件重命名建议（金额-购买方格式）:")
        rename_ops = rename_invoice_files(invoices, rename=args.rename)
        for op in rename_ops[:20]:  # 仅显示前20条
            print(f"  {op}")
        if len(rename_ops) > 20:
            print(f"  ... 还有 {len(rename_ops) - 20} 条")

        if args.rename:
            print(f"\n✅ 已重命名 {sum(1 for op in rename_ops if op.startswith('✓'))} 份文件")

    print("\n" + "=" * 80)
    print("✨ 处理完成！")
    return 0


if __name__ == "__main__":
    sys.exit(main())
